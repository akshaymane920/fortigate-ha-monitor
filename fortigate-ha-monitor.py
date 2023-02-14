import json
from openpyxl.styles import PatternFill
from pyFortiManagerAPI import FortiManager
import getpass
from openpyxl import Workbook

# Step1: Login into the FortiManager using API
fortiManager_ip = input("FortiManager IP: ")
username = input("Username: ")
password = getpass.getpass("Password: ")
fmg = FortiManager(host=fortiManager_ip, username=username, password=password, verify=False)

# Step2: Get All FortiGate devices managed by FortiManager
get_devices = fmg.get_devices()["result"][0]["data"]
# print(json.dumps(get_devices, indent=4))


# Step3: Fetch the HA status of all devices
result = []
for device in get_devices:
    template = {}
    template.update({"name": device["name"]})
    if device["ha_mode"] == 0:
        status = False if device["conn_status"] == 2 else True
        template.update(
            {"ha_mode": False, "members": [{"id": 1, "member1": device["name"], "status": status}]})
    else:
        template.update({"ha_mode": True})
        members = []
        _id = 1
        for member in device["ha_slave"]:
            status = False if member["status"] == 2 else True
            members.append({"id": _id, "member": member["name"], "status": status})
            _id += 1
        template.update({"members": members})
    result.append(template)
print(json.dumps(result, indent=4))

# Step4: Create an excel report (Optional)
excelReport = Workbook()
ROW = 1
sheet = excelReport.create_sheet(title="HA Status")
sheet.cell(row=ROW, column=1, value="name")
sheet.cell(row=ROW, column=2, value="in ha?")
sheet.cell(row=ROW, column=3, value="member1")
sheet.cell(row=ROW, column=4, value="member2")

ROW = 2
for device in result:
    sheet.cell(row=ROW, column=1, value=device["name"])
    if device["ha_mode"]:
        sheet.cell(row=ROW, column=2, value="yes")
        COL = 3
        for member in device["members"]:
            cell = sheet.cell(row=ROW, column=COL, value=member["member"])
            if member["status"] == 1:
                cell.fill = PatternFill(fill_type="solid", fgColor="E0FFD1")
            else:
                cell.fill = PatternFill(fill_type="solid", fgColor="FFD1D1")
            COL += 1
    else:
        sheet.cell(row=ROW, column=2, value="no")
        cell = sheet.cell(row=ROW, column=3, value=device["name"])
        if device["members"][0]["status"] == 1:
            cell.fill = PatternFill(fill_type="solid", fgColor="E0FFD1")
        else:
            cell.fill = PatternFill(fill_type="solid", fgColor="FFD1D1")
    ROW += 1
excelReport.remove(excelReport['Sheet'])
excelReport.save("ha_status.xlsx")
